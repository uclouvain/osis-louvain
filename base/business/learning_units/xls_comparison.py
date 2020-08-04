##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2019 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
from decimal import Decimal

from django.utils.translation import gettext_lazy as _
from openpyxl.utils import get_column_letter

from base.business import learning_unit_year_with_context
from base.business.entity import build_entity_container_prefetch
from base.business.learning_unit import get_organization_from_learning_unit_year
from base.business.learning_unit_year_with_context import append_latest_entities, append_components, \
    get_learning_component_prefetch
from base.business.learning_units.comparison import get_partims_as_str
from base.business.proposal_xls import BLANK_VALUE, XLS_DESCRIPTION_COMPARISON, XLS_COMPARISON_FILENAME, \
    COMPARISON_PROPOSAL_TITLES, COMPARISON_WORKSHEET_TITLE, basic_titles, components_titles
from base.business.utils.convert import volume_format
from base.business.xls import get_name_or_username
from base.enums.component_detail import VOLUME_TOTAL, VOLUME_Q1, VOLUME_Q2, PLANNED_CLASSES, \
    VOLUME_REQUIREMENT_ENTITY, VOLUME_ADDITIONAL_REQUIREMENT_ENTITY_1, VOLUME_ADDITIONAL_REQUIREMENT_ENTITY_2, \
    VOLUME_TOTAL_REQUIREMENT_ENTITIES, REAL_CLASSES, VOLUME_GLOBAL
from base.models.academic_year import find_academic_year_by_id
from base.models.campus import find_by_id as find_campus_by_id
from base.models.entity import find_by_id
from base.models.enums import entity_container_year_link_type as entity_types, vacant_declaration_type, \
    attribution_procedure
from base.models.enums import learning_component_year_type
from base.models.enums.learning_component_year_type import LECTURING, PRACTICAL_EXERCISES
from base.models.enums.learning_container_year_types import LearningContainerYearType
from base.models.enums.learning_unit_year_periodicity import PERIODICITY_TYPES
from base.models.learning_unit_year import LearningUnitYear, get_by_id
from osis_common.document import xls_build
from reference.models.language import find_by_id as find_language_by_id

EMPTY_VALUE = ''
DATE_FORMAT = '%d-%m-%Y'
DATE_TIME_FORMAT = '%d-%m-%Y %H:%M'
DESC = "desc"
WORKSHEET_TITLE = 'learning_units_comparison'
XLS_FILENAME = 'learning_units_comparison'
XLS_DESCRIPTION = _("Comparison of learning units")

ACRONYM_COL_NUMBER = 0
ACADEMIC_COL_NUMBER = 1
CELLS_MODIFIED_NO_BORDER = 'modifications'
CELLS_TOP_BORDER = 'border_not_modified'
DATA = 'data'


def learning_unit_titles():
    return basic_titles() + components_titles()


def create_xls_comparison(user, learning_unit_years, filters, academic_yr_comparison):
    working_sheets_data = []
    cells_modified_with_green_font = []
    cells_with_top_border = []

    if learning_unit_years:
        luys_for_2_years = _get_learning_unit_yrs_on_2_different_years(academic_yr_comparison, learning_unit_years)
        data = prepare_xls_content(luys_for_2_years)
        working_sheets_data = data.get('data')
        cells_modified_with_green_font = data.get(CELLS_MODIFIED_NO_BORDER)
        cells_with_top_border = data.get(CELLS_TOP_BORDER)
    parameters = {
        xls_build.DESCRIPTION: XLS_DESCRIPTION,
        xls_build.USER: get_name_or_username(user),
        xls_build.FILENAME: XLS_FILENAME,
        xls_build.HEADER_TITLES: learning_unit_titles(),
        xls_build.WS_TITLE: WORKSHEET_TITLE,
    }
    dict_styled_cells = {}

    if cells_modified_with_green_font:
        parameters[xls_build.FONT_CELLS] = parameters[xls_build.FONT_CELLS].update(
            {xls_build.STYLE_MODIFIED: cells_modified_with_green_font}
        )

    if cells_with_top_border:
        parameters[xls_build.BORDER_CELLS] = dict_styled_cells.update({xls_build.BORDER_TOP: cells_with_top_border})

    if dict_styled_cells:
        parameters[xls_build.STYLED_CELLS] = dict_styled_cells

    return xls_build.generate_xls(xls_build.prepare_xls_parameters_list(working_sheets_data, parameters), filters)


def _get_learning_unit_yrs_on_2_different_years(academic_yr_comparison, learning_unit_years):
    learning_unit_years = LearningUnitYear.objects.filter(
        learning_unit__in=(_get_learning_units(learning_unit_years)),
        academic_year__year__in=(
            learning_unit_years[0].academic_year.year,
            academic_yr_comparison)
    ).select_related(
        'academic_year',
        'learning_container_year',
        'learning_container_year__academic_year'
    ).prefetch_related(
        get_learning_component_prefetch()
    ).prefetch_related(
        build_entity_container_prefetch(entity_types.ALLOCATION_ENTITY),
        build_entity_container_prefetch(entity_types.REQUIREMENT_ENTITY),
        build_entity_container_prefetch(entity_types.ADDITIONAL_REQUIREMENT_ENTITY_1),
        build_entity_container_prefetch(entity_types.ADDITIONAL_REQUIREMENT_ENTITY_2),
    ).order_by('learning_unit', 'academic_year__year')
    [append_latest_entities(learning_unit) for learning_unit in learning_unit_years]
    [append_components(learning_unit) for learning_unit in learning_unit_years]
    return learning_unit_years


def _get_learning_units(learning_unit_years):
    return list(set([l.learning_unit for l in learning_unit_years]))


def prepare_xls_content(learning_unit_yrs):
    data = []
    learning_unit = None
    first_data = None
    modified_cells_no_border = []
    top_border = []
    for line_index, l_u_yr in enumerate(learning_unit_yrs, start=1):

        if learning_unit is None:
            learning_unit = l_u_yr.learning_unit
            new_line = True
        else:
            if learning_unit == l_u_yr.learning_unit:
                new_line = False
            else:
                learning_unit = l_u_yr.learning_unit
                new_line = True
        luy_data = extract_xls_data_from_learning_unit(l_u_yr, new_line, first_data)
        if new_line:
            first_data = luy_data
            top_border.extend(get_border_columns(line_index + 1))
        else:
            modified_cells_no_border.extend(
                _check_changes_other_than_code_and_year(first_data, luy_data, line_index + 1))
            first_data = None
        data.append(luy_data)

    return {
        DATA: data,
        CELLS_TOP_BORDER: top_border or None,
        CELLS_MODIFIED_NO_BORDER: modified_cells_no_border or None,
    }


def extract_xls_data_from_learning_unit(learning_unit_yr, new_line, first_data):
    data = _get_data(learning_unit_yr, new_line, first_data)
    data.extend(_component_data(learning_unit_yr.components, learning_component_year_type.LECTURING))
    data.extend(_component_data(learning_unit_yr.components, learning_component_year_type.PRACTICAL_EXERCISES))
    return data


def translate_status(value):
    if value:
        return _('Active')
    else:
        return _('Inactive')


def _component_data(components, learning_component_yr_type):
    if components:
        for component in components:
            if component.type == learning_component_yr_type:
                return _get_volumes(component, components)
    return [EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE,
            EMPTY_VALUE, EMPTY_VALUE]


def _get_data(learning_unit_yr, new_line, first_data, partims=True, proposal_comparison=False):
    organization = get_organization_from_learning_unit_year(learning_unit_yr)
    if proposal_comparison:
        academic_year = _format_academic_year(
            learning_unit_yr.academic_year.name,
            learning_unit_yr.learning_unit.end_year.name if learning_unit_yr.learning_unit.end_year else None
        )
    else:
        academic_year = learning_unit_yr.academic_year.name

    data = [
        _get_acronym(learning_unit_yr, new_line, first_data),
        academic_year,
        learning_unit_yr.learning_container_year.get_container_type_display()
        if learning_unit_yr.learning_container_year.container_type else BLANK_VALUE,
        translate_status(learning_unit_yr.status),
        learning_unit_yr.get_subtype_display() if learning_unit_yr.subtype else BLANK_VALUE,
        learning_unit_yr.get_internship_subtype_display() if learning_unit_yr.internship_subtype else BLANK_VALUE,
        volume_format(learning_unit_yr.credits) or BLANK_VALUE,
        learning_unit_yr.language.name or BLANK_VALUE,
        learning_unit_yr.get_periodicity_display() if learning_unit_yr.periodicity else BLANK_VALUE,
        get_translation(learning_unit_yr.quadrimester),
        get_translation(learning_unit_yr.session),
        get_representing_string(learning_unit_yr.learning_container_year.common_title),
        get_representing_string(learning_unit_yr.specific_title),
        get_representing_string(learning_unit_yr.learning_container_year.common_title_english),
        get_representing_string(learning_unit_yr.specific_title_english),
        _get_entity_to_display(learning_unit_yr.entities.get(entity_types.REQUIREMENT_ENTITY)),
        _get_entity_to_display(learning_unit_yr.entities.get(entity_types.ALLOCATION_ENTITY)),
        _get_entity_to_display(learning_unit_yr.entities.get(entity_types.ADDITIONAL_REQUIREMENT_ENTITY_1)),
        _get_entity_to_display(learning_unit_yr.entities.get(entity_types.ADDITIONAL_REQUIREMENT_ENTITY_2)),
        _('Yes') if learning_unit_yr.professional_integration else _('No'),
        organization.name if organization else BLANK_VALUE,
        learning_unit_yr.campus or BLANK_VALUE]
    if partims:
        data.append(get_partims_as_str(learning_unit_yr.get_partims_related()))
    data.extend(
        [
            get_representing_string(learning_unit_yr.learning_unit.faculty_remark),
            get_representing_string(learning_unit_yr.learning_unit.other_remark),
            _('Yes') if learning_unit_yr.learning_container_year.team else _('No'),
            _('Yes') if learning_unit_yr.learning_container_year.is_vacant else _('No'),
            get_representing_string(learning_unit_yr.learning_container_year.get_type_declaration_vacant_display()),
            get_representing_string(learning_unit_yr.get_attribution_procedure_display()),
        ]
    )

    return data


def _get_acronym(learning_unit_yr, new_line, first_data):
    if first_data:
        acronym = EMPTY_VALUE
        if new_line:
            acronym = learning_unit_yr.acronym
        else:
            if learning_unit_yr.acronym != first_data[ACRONYM_COL_NUMBER]:
                acronym = learning_unit_yr.acronym
        return acronym
    else:
        return learning_unit_yr.acronym


def _get_volumes(component, components):
    volumes = components[component]
    return [
        component.acronym if component.acronym else EMPTY_VALUE,
        volumes.get(VOLUME_Q1, EMPTY_VALUE),
        volumes.get(VOLUME_Q2, EMPTY_VALUE),
        volumes.get(VOLUME_TOTAL, EMPTY_VALUE),
        component.real_classes if component.real_classes else EMPTY_VALUE,
        component.planned_classes if component.planned_classes else EMPTY_VALUE,
        volumes.get(VOLUME_GLOBAL, '0'),
        volumes.get(VOLUME_REQUIREMENT_ENTITY, EMPTY_VALUE),
        volumes.get(VOLUME_ADDITIONAL_REQUIREMENT_ENTITY_1, EMPTY_VALUE),
        volumes.get(VOLUME_ADDITIONAL_REQUIREMENT_ENTITY_2, EMPTY_VALUE)
    ]


def get_translation(value):
    return str(_(value)) if value else BLANK_VALUE


def _get_entity_to_display(entity):
    return entity.acronym if entity else BLANK_VALUE


def _check_changes_other_than_code_and_year(first_data, second_data, line_index):
    modifications = []
    for col_index, obj in enumerate(first_data):
        if col_index == ACRONYM_COL_NUMBER and second_data[ACRONYM_COL_NUMBER] != EMPTY_VALUE:
            modifications.append('{}{}'.format(get_column_letter(col_index + 1), line_index))
        else:
            if obj != second_data[col_index] and col_index != ACADEMIC_COL_NUMBER:
                modifications.append('{}{}'.format(get_column_letter(col_index + 1), line_index))

    return modifications


def get_border_columns(line):
    style = []
    for col_index, obj in enumerate(learning_unit_titles(), start=1):
        style.append('{}{}'.format(get_column_letter(col_index), line))
    return style


def _get_component_data_by_type(component):
    if component:
        return [
            get_representing_string(component.get(VOLUME_Q1)),
            get_representing_string(component.get(VOLUME_Q2)),
            get_representing_string(component.get(VOLUME_TOTAL)),
            get_representing_string(component.get(REAL_CLASSES)),
            get_representing_string(component.get(PLANNED_CLASSES)),
            get_representing_string(component.get(VOLUME_TOTAL_REQUIREMENT_ENTITIES)),
            get_representing_string(component.get(VOLUME_REQUIREMENT_ENTITY)),
            get_representing_string(component.get(VOLUME_ADDITIONAL_REQUIREMENT_ENTITY_1)),
            get_representing_string(component.get(VOLUME_ADDITIONAL_REQUIREMENT_ENTITY_2))
        ]
    else:
        return [BLANK_VALUE, BLANK_VALUE, BLANK_VALUE, BLANK_VALUE, BLANK_VALUE, BLANK_VALUE, BLANK_VALUE,
                BLANK_VALUE, BLANK_VALUE]


def prepare_xls_content_for_comparison(luy_with_proposals):
    line_index = 1
    data = []
    top_border = []
    modified_cells_no_border = []
    for luy_with_proposal in luy_with_proposals:
        top_border.extend(get_border_columns(line_index))
        data_proposal = _get_proposal_data(luy_with_proposal)
        data.append(data_proposal)

        proposal = luy_with_proposal.proposallearningunit
        initial_luy_data = proposal.initial_data

        if initial_luy_data and initial_luy_data.get('learning_unit'):
            initial_data = _get_data_from_initial_data(luy_with_proposal.proposallearningunit.initial_data, True)
            data.append(initial_data)
            modified_cells_no_border.extend(
                _check_changes(initial_data,
                               data_proposal,
                               line_index + 1))
            line_index += 2
        else:
            line_index += 1

    return {
        DATA: data,
        CELLS_TOP_BORDER: top_border or None,
        CELLS_MODIFIED_NO_BORDER: modified_cells_no_border or None,
    }


def _get_data_from_initial_data(initial_data, proposal_comparison=False):
    luy_initial = initial_data.get('learning_unit_year', {})
    lcy_initial = initial_data.get('learning_container_year', {})
    lu_initial = initial_data.get('learning_unit', {})

    if luy_initial.get('id'):
        learning_unit_yr = get_by_id(luy_initial.get('id'))
    else:
        learning_unit_yr = None

    requirement_entity = find_by_id(lcy_initial.get('requirement_entity'))
    allocation_entity = find_by_id(lcy_initial.get('allocation_entity'))
    add1_requirement_entity = find_by_id(lcy_initial.get('additional_entity_1'))
    add2_requirement_entity = find_by_id(lcy_initial.get('additional_entity_2'))
    campus = find_campus_by_id(luy_initial.get('campus'))

    organization = None
    if learning_unit_yr:
        organization = get_organization_from_learning_unit_year(learning_unit_yr)
    language = find_language_by_id(luy_initial.get('language'))

    if proposal_comparison:
        academic_year = _format_academic_year(learning_unit_yr.academic_year.name,
                                              find_academic_year_by_id(lu_initial.get('end_year'))
                                              if lu_initial.get('end_year') else None)
    else:
        academic_year = learning_unit_yr.academic_year.name

    data = [
        str(_('Initial data')),
        luy_initial.get('acronym', ''),
        academic_year,
        dict(LearningContainerYearType.choices())[lcy_initial.get('container_type')] if
        lcy_initial.get('container_type') else BLANK_VALUE,
        translate_status(luy_initial.get('status')),
        learning_unit_yr.get_subtype_display()
        if learning_unit_yr and learning_unit_yr.get_subtype_display() else BLANK_VALUE,
        get_translation(luy_initial.get('internship_subtype')),
        volume_format(Decimal(luy_initial['credits'])) if luy_initial.get('credits') else BLANK_VALUE,
        language.name if language else BLANK_VALUE,
        dict(PERIODICITY_TYPES)[luy_initial['periodicity']] if luy_initial.get('periodicity') else BLANK_VALUE,
        get_translation(luy_initial.get('quadrimester')),
        get_translation(luy_initial.get('session')),
        get_representing_string(lcy_initial.get('common_title')),
        get_representing_string(luy_initial.get('specific_title')),
        get_representing_string(lcy_initial.get('common_title_english')),
        get_representing_string(luy_initial.get('specific_title_english')),
        requirement_entity.most_recent_acronym if requirement_entity else BLANK_VALUE,
        allocation_entity.most_recent_acronym if allocation_entity else BLANK_VALUE,
        add1_requirement_entity.most_recent_acronym if add1_requirement_entity else BLANK_VALUE,
        add2_requirement_entity.most_recent_acronym if add2_requirement_entity else BLANK_VALUE,
        _('Yes') if luy_initial.get('professional_integration') else _('No'),
        organization.name if organization else BLANK_VALUE,
        campus if campus else BLANK_VALUE,
        get_representing_string(lu_initial.get('faculty_remark')),
        get_representing_string(lu_initial.get('other_remark')),
        _('Yes') if lcy_initial.get('team') else _('No'),
        _('Yes') if lcy_initial.get('is_vacant') else _('No'),
        dict(vacant_declaration_type.DECLARATION_TYPE)[lcy_initial.get('type_declaration_vacant')] if lcy_initial.get(
            'type_declaration_vacant') else BLANK_VALUE,
        dict(attribution_procedure.ATTRIBUTION_PROCEDURES)[luy_initial.get('attribution_procedure')] if luy_initial.get(
            'attribution_procedure') else BLANK_VALUE,
    ]
    return _get_data_from_components_initial_data(data, initial_data)


def _check_changes(initial_data, proposal_data, line_index):
    modifications = []
    for col_index, obj in enumerate(initial_data[2:]):
        if str(obj) != str(proposal_data[col_index + 2]):
            modifications.append('{}{}'.format(get_column_letter(col_index + 2 + 1), line_index))
    return modifications


def get_representing_string(value):
    return value or BLANK_VALUE


def create_xls_proposal_comparison(user, learning_units_with_proposal, filters):
    data = prepare_xls_content_for_comparison(learning_units_with_proposal)

    working_sheets_data = data.get('data')
    cells_modified_with_green_font = data.get(CELLS_MODIFIED_NO_BORDER)
    cells_with_top_border = data.get(CELLS_TOP_BORDER)

    parameters = {
        xls_build.DESCRIPTION: XLS_DESCRIPTION_COMPARISON,
        xls_build.USER: get_name_or_username(user),
        xls_build.FILENAME: XLS_COMPARISON_FILENAME,
        xls_build.HEADER_TITLES: COMPARISON_PROPOSAL_TITLES,
        xls_build.WS_TITLE: COMPARISON_WORKSHEET_TITLE,
    }
    dict_styled_cells = {}
    if cells_modified_with_green_font:
        parameters[xls_build.FONT_CELLS] = parameters[xls_build.FONT_CELLS].update(
            {xls_build.STYLE_MODIFIED: cells_modified_with_green_font}
        )

    if cells_with_top_border:
        dict_styled_cells[xls_build.BORDER_BOTTOM] = cells_with_top_border
    if dict_styled_cells:
        parameters[xls_build.STYLED_CELLS] = dict_styled_cells
    return xls_build.generate_xls(xls_build.prepare_xls_parameters_list(working_sheets_data, parameters), filters)


def _get_basic_components(learning_unit_yr):
    learning_unit_yr = find_learning_unit_yr_with_components_data(learning_unit_yr)
    components = []
    components_values = []
    for key, value in learning_unit_yr.components.items():
        components.append(key)
        components_values.append(value)

    practical_component = None
    lecturing_component = None
    for index, component in enumerate(components):
        if not practical_component and component.type == PRACTICAL_EXERCISES:
            practical_component = _build_component(component.real_classes, components_values, index)

        if not lecturing_component and component.type == LECTURING:
            lecturing_component = _build_component(component.real_classes, components_values, index)
    return {PRACTICAL_EXERCISES: practical_component, LECTURING: lecturing_component}


def _build_component(real_classes, components_values, index):
    a_component = components_values[index]
    a_component['REAL_CLASSES'] = real_classes
    return a_component


def _get_components_data(learning_unit_yr):
    components_data_dict = _get_basic_components(learning_unit_yr)
    return \
        _get_component_data_by_type(components_data_dict.get(LECTURING)) + \
        _get_component_data_by_type(components_data_dict.get(PRACTICAL_EXERCISES))


def _get_proposal_data(learning_unit_yr):
    data_proposal = [_('Proposal')] + _get_data(learning_unit_yr, False, None, False, True)
    data_proposal.extend(_get_components_data(learning_unit_yr))
    return data_proposal


def find_learning_unit_yr_with_components_data(learning_unit_yr):
    learning_unit_yrs = learning_unit_year_with_context.get_with_context(
        learning_container_year_id=learning_unit_yr.learning_container_year.id
    )
    if learning_unit_yrs:
        learning_unit_yr = next(luy for luy in learning_unit_yrs if luy.id == learning_unit_yr.id)

    return learning_unit_yr


def _get_data_from_components_initial_data(data_without_components, initial_data):
    data = data_without_components
    volumes = initial_data.get('volumes')
    if volumes:
        data = data + _get_component_data_by_type(volumes.get('PM'))
        data = data + _get_component_data_by_type(volumes.get('PP'))
    return data


def _format_academic_year(start_year, end_year):
    return "{}{}".format(start_year,
                         "   ({} {})".format(_('End').lower(), end_year if end_year else '-'))
