##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2021 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
from collections import defaultdict
from typing import List, Dict

from django.db.models import QuerySet
from django.db.models import Subquery, OuterRef
from django.db.models.expressions import RawSQL
from django.template.defaultfilters import yesno
from django.utils.translation import gettext_lazy as _
from openpyxl.styles import Alignment, PatternFill, Color, Font
from openpyxl.utils import get_column_letter

from attribution.business import attribution_charge_new
from attribution.models.enums.function import Functions
from base.business.xls import get_name_or_username, _get_all_columns_reference
from base.models.enums.learning_component_year_type import LECTURING, PRACTICAL_EXERCISES
from base.models.enums.proposal_type import ProposalType
from base.models.group_element_year import GroupElementYear
from base.models.learning_component_year import LearningComponentYear
from base.models.learning_unit_year import SQL_RECURSIVE_QUERY_EDUCATION_GROUP_TO_CLOSEST_TRAININGS, LearningUnitYear
from base.models.person import Person
from osis_common.document import xls_build
from program_management.ddd.domain.program_tree_version import ProgramTreeVersionIdentity, version_label

XLS_DESCRIPTION = _('Learning units list')
XLS_FILENAME = _('LearningUnitsList')
WORKSHEET_TITLE = _('Learning units list')

TRANSFORMATION_AND_MODIFICATION_COLOR = Color('808000')
TRANSFORMATION_COLOR = Color('ff6600')
SUPPRESSION_COLOR = Color('ff0000')
MODIFICATION_COLOR = Color('0000ff')
CREATION_COLOR = Color('008000')
DEFAULT_LEGEND_FILLS = {
    PatternFill(patternType='solid', fgColor=CREATION_COLOR): ['A2'],
    PatternFill(patternType='solid', fgColor=MODIFICATION_COLOR): ['A3'],
    PatternFill(patternType='solid', fgColor=SUPPRESSION_COLOR): ['A4'],
    PatternFill(patternType='solid', fgColor=TRANSFORMATION_COLOR): ['A5'],
    PatternFill(patternType='solid', fgColor=TRANSFORMATION_AND_MODIFICATION_COLOR): ['A6'],
}
BOLD_FONT = Font(bold=True)
SPACES = '  '
HEADER_TEACHERS = [
    str(_('List of teachers')),
    str(_('List of teachers (email)'))
]
HEADER_PROGRAMS = _('Trainings')
PROPOSAL_LINE_STYLES = {
    ProposalType.CREATION.name: Font(color=CREATION_COLOR),
    ProposalType.MODIFICATION.name: Font(color=MODIFICATION_COLOR),
    ProposalType.SUPPRESSION.name: Font(color=SUPPRESSION_COLOR),
    ProposalType.TRANSFORMATION.name: Font(color=TRANSFORMATION_COLOR),
    ProposalType.TRANSFORMATION_AND_MODIFICATION.name: Font(color=TRANSFORMATION_AND_MODIFICATION_COLOR),
}
WRAP_TEXT_ALIGNMENT = Alignment(wrapText=True, vertical="top")
WITH_ATTRIBUTIONS = 'with_attributions'
WITH_GRP = 'with_grp'


def learning_unit_titles_part1() -> List[str]:
    return [
        str(_('Code')),
        str(_('Ac yr.')),
        str(_('Title')),
        str(_('Type')),
        str(_('Subtype')),
        "{} ({})".format(_('Req. Entity'), _('fac. level')),
        str(_('Proposal type')),
        str(_('Proposal status')),
        str(_('Credits')),
        str(_('Alloc. Ent.')),
        str(_('Title in English')),
    ]


def prepare_xls_content(learning_unit_years: QuerySet,
                        is_external_ue_list: bool,
                        with_grp=False,
                        with_attributions=False) -> List:
    qs = annotate_qs(learning_unit_years)

    if with_grp:
        qs = qs.annotate(
            closest_trainings=RawSQL(SQL_RECURSIVE_QUERY_EDUCATION_GROUP_TO_CLOSEST_TRAININGS, ())
        ).prefetch_related('element')

    result = []

    for learning_unit_yr in qs:
        lu_data_part1 = get_data_part1(learning_unit_yr, is_external_ue_list)
        lu_data_part2 = get_data_part2(learning_unit_yr, with_attributions)

        if with_grp:
            lu_data_part2.append(_add_training_data(learning_unit_yr))

        lu_data_part1.extend(lu_data_part2)
        if is_external_ue_list:
            lu_data_part1.extend(_get_external_ue_data(learning_unit_yr))
        result.append(lu_data_part1)

    return result


def annotate_qs(learning_unit_years: QuerySet) -> QuerySet:
    """ Fetch directly in the queryset all volumes data."""

    subquery_component = LearningComponentYear.objects.filter(
        learning_unit_year__in=OuterRef('pk')
    )
    subquery_component_pm = subquery_component.filter(
        type=LECTURING
    )
    subquery_component_pp = subquery_component.filter(
        type=PRACTICAL_EXERCISES
    )

    return learning_unit_years.annotate(
        pm_vol_q1=Subquery(subquery_component_pm.values('hourly_volume_partial_q1')[:1]),
        pm_vol_q2=Subquery(subquery_component_pm.values('hourly_volume_partial_q2')[:1]),
        pm_vol_tot=Subquery(subquery_component_pm.values('hourly_volume_total_annual')[:1]),
        pm_classes=Subquery(subquery_component_pm.values('planned_classes')[:1]),
        pp_vol_q1=Subquery(subquery_component_pp.values('hourly_volume_partial_q1')[:1]),
        pp_vol_q2=Subquery(subquery_component_pp.values('hourly_volume_partial_q2')[:1]),
        pp_vol_tot=Subquery(subquery_component_pp.values('hourly_volume_total_annual')[:1]),
        pp_classes=Subquery(subquery_component_pp.values('planned_classes')[:1])
    )


def create_xls_with_parameters(user, learning_units, filters, extra_configuration, is_external_ue_list):
    with_grp = extra_configuration.get(WITH_GRP)
    with_attributions = extra_configuration.get(WITH_ATTRIBUTIONS)
    titles_part1 = _prepare_titles(is_external_ue_list, with_attributions, with_grp)

    working_sheet_data = prepare_xls_content(
        learning_units,
        is_external_ue_list,
        with_grp,
        with_attributions
    )
    ws_data = xls_build.prepare_xls_parameters_list(
        working_sheet_data,
        _get_parameters_configurable_list(learning_units, titles_part1, user)
    )
    working_sheets_data = [ws_data.get(xls_build.WORKSHEETS_DATA)[0]]
    if not is_external_ue_list:
        working_sheets_data.append(prepare_proposal_legend_ws_data())
    ws_data.update({xls_build.WORKSHEETS_DATA: working_sheets_data})
    return xls_build.generate_xls(ws_data, filters)


def _get_parameters_configurable_list(learning_units, titles, user) -> dict:
    parameters = {
        xls_build.DESCRIPTION: XLS_DESCRIPTION,
        xls_build.USER: get_name_or_username(user),
        xls_build.FILENAME: XLS_FILENAME,
        xls_build.HEADER_TITLES: titles,
        xls_build.WS_TITLE: WORKSHEET_TITLE,
        xls_build.ALIGN_CELLS: {
            WRAP_TEXT_ALIGNMENT: _get_wrapped_cells(
                learning_units,
                _get_col_letter(titles, HEADER_PROGRAMS),
                _get_col_letter(titles, HEADER_TEACHERS)
            )
        },
        xls_build.FONT_ROWS: _get_font_rows(learning_units),
    }
    return parameters


def get_significant_volume(volume):
    if volume and volume > 0:
        return volume
    return ''


def prepare_proposal_legend_ws_data() -> dict:
    return {
        xls_build.HEADER_TITLES_KEY: [str(_('Legend'))],
        xls_build.CONTENT_KEY: [
            [SPACES, _('Proposal of creation')],
            [SPACES, _('Proposal for modification')],
            [SPACES, _('Suppression proposal')],
            [SPACES, _('Transformation proposal')],
            [SPACES, _('Transformation/modification proposal')],
        ],
        xls_build.WORKSHEET_TITLE_KEY: _('Legend'),
        xls_build.STYLED_CELLS:
            DEFAULT_LEGEND_FILLS
    }


def _get_wrapped_cells(learning_units, programs_col_letter, teachers_col_letter):
    dict_wrapped_styled_cells = []

    for idx, luy in enumerate(learning_units, start=2):
        if teachers_col_letter:
            for teacher_col_letter in teachers_col_letter:
                dict_wrapped_styled_cells.append("{}{}".format(teacher_col_letter, idx))
        if programs_col_letter:
            dict_wrapped_styled_cells.append("{}{}".format(programs_col_letter, idx))

    return dict_wrapped_styled_cells


def _get_font_rows(learning_units):
    colored_cells = defaultdict(list)
    for idx, luy in enumerate(learning_units, start=1):
        if getattr(luy, "proposallearningunit", None):
            colored_cells[PROPOSAL_LINE_STYLES.get(luy.proposallearningunit.type)].append(idx)
    colored_cells.update({BOLD_FONT: [0]})
    return colored_cells


def _get_attribution_line(an_attribution):
    return "{} - {} : {} - {} : {} - {} : {} - {} : {} - {} : {} - {} : {} ".format(
        an_attribution.get('person'),
        _('Function'),
        Functions[an_attribution['function']].value if 'function' in an_attribution else '',
        _('Substitute'),
        an_attribution.get('substitute') if an_attribution.get('substitute') else '',
        _('Beg. of attribution'),
        an_attribution.get('start_year'),
        _('Attribution duration'),
        an_attribution.get('duration'),
        _('Attrib. vol1'),
        an_attribution.get('LECTURING'),
        _('Attrib. vol2'),
        an_attribution.get('PRACTICAL_EXERCISES'),
    )


def _get_col_letter(titles, title_search):
    for idx, title in enumerate(titles, start=1):
        if title == title_search:
            return get_column_letter(idx)
    return None


def _add_training_data(learning_unit_yr: LearningUnitYear) -> str:
    return ("\n".join([
        _concatenate_training_data(learning_unit_yr, group_element_year)
        for group_element_year in learning_unit_yr.element.children_elements.all()
    ])).strip() if hasattr(learning_unit_yr, "element") else ''


def _concatenate_training_data(learning_unit_year: LearningUnitYear, group_element_year: GroupElementYear) -> str:
    concatenated_string = ''
    if not learning_unit_year.closest_trainings or group_element_year.parent_element.group_year is None:
        return concatenated_string

    partial_acronym = group_element_year.parent_element.group_year.partial_acronym or ''
    credits = group_element_year.relative_credits \
        if group_element_year.relative_credits else group_element_year.child_element.learning_unit_year.credits
    leaf_credits = "{0:.2f}".format(credits) if credits else '-'
    nb_parents = '-' if len(learning_unit_year.closest_trainings) > 0 else ''

    for training in learning_unit_year.closest_trainings:
        if training['gs_origin'] == group_element_year.pk:
            training_string = "{} ({}) {} {} - {}\n".format(
                partial_acronym,
                leaf_credits,
                nb_parents,
                acronym_with_version_label(
                    training['acronym'], training['transition_name'], training['version_name']
                ),
                title_with_version_title(training['title_fr'], training['version_title_fr']),
            )
            concatenated_string += training_string

    return concatenated_string


def title_with_version_title(title_fr: str, version_title_fr: str) -> str:
    return title_fr + (' [{}]'.format(version_title_fr) if version_title_fr else '')


def get_data_part2(learning_unit_yr: LearningUnitYear, with_attributions: bool) -> List[str]:
    lu_data_part2 = []
    if with_attributions:
        teachers = _get_teachers(learning_unit_yr)
        lu_data_part2.append(';'.join(_build_complete_name(person) for person in teachers))
        lu_data_part2.append(';'.join(person.email if person.email else '-' for person in teachers))

    lu_data_part2.append(learning_unit_yr.get_periodicity_display())
    lu_data_part2.append(yesno(learning_unit_yr.status))
    lu_data_part2.extend(volume_information(learning_unit_yr))
    lu_data_part2.extend([
        learning_unit_yr.get_quadrimester_display() or '',
        learning_unit_yr.get_session_display() or '',
        learning_unit_yr.language or "",
    ])
    return lu_data_part2


def get_data_part1(learning_unit_yr: LearningUnitYear, is_external_ue_list: bool) -> List[str]:
    proposal = getattr(learning_unit_yr, "proposallearningunit", None)
    requirement_acronym = learning_unit_yr.entity_requirement
    allocation_acronym = learning_unit_yr.entity_allocation

    lu_common_data_part1 = [
        learning_unit_yr.acronym,
        learning_unit_yr.academic_year.name,
        learning_unit_yr.complete_title,
        learning_unit_yr.get_container_type_display(),
        learning_unit_yr.get_subtype_display(),
        requirement_acronym,
        ]
    if is_external_ue_list:
        lu_proposal_data = []
    else:
        lu_proposal_data = [
            proposal.get_type_display() if proposal else '',
            proposal.get_state_display() if proposal else '',
        ]

    lu_common_data_part2 = [
        learning_unit_yr.credits,
        allocation_acronym,
        learning_unit_yr.complete_title_english,
    ]
    return lu_common_data_part1 + lu_proposal_data + lu_common_data_part2


def learning_unit_titles_part2() -> List[str]:
    return [
        str(_('Periodicity')),
        str(_('Active')),
        "{} - {}".format(_('Lecturing vol.'), _('Annual')),
        "{} - {}".format(_('Lecturing vol.'), _('1st quadri')),
        "{} - {}".format(_('Lecturing vol.'), _('2nd quadri')),
        "{}".format(_('Lecturing planned classes')),
        "{} - {}".format(_('Practical vol.'), _('Annual')),
        "{} - {}".format(_('Practical vol.'), _('1st quadri')),
        "{} - {}".format(_('Practical vol.'), _('2nd quadri')),
        "{}".format(_('Practical planned classes')),
        str(_('Quadrimester')),
        str(_('Session derogation')),
        str(_('Language')),
    ]


def learning_unit_titles_part_1(display_proposal: bool) -> List[str]:
    common_titles_part1 = [
        str(_('Code')),
        str(_('Ac yr.')),
        str(_('Title')),
        str(_('Type')),
        str(_('Subtype')),
        str(_('Req. Entity')),
    ]
    if display_proposal:
        proposal_titles = [
            str(_('Proposal type')),
            str(_('Proposal status')),
            ]
    else:
        proposal_titles = []
    common_title_part2 = [
        str(_('Credits')),
        str(_('Alloc. Ent.')),
        str(_('Title in English')),
    ]
    return common_titles_part1 + proposal_titles + common_title_part2


def prepare_ue_xls_content(found_learning_units):
    return [extract_xls_data_from_learning_unit(lu) for lu in found_learning_units]


def extract_xls_data_from_learning_unit(learning_unit_yr: LearningUnitYear) -> List[str]:
    return [
        learning_unit_yr.academic_year.name, learning_unit_yr.acronym, learning_unit_yr.complete_title,
        xls_build.translate(learning_unit_yr.learning_container_year.container_type)
        # FIXME Condition to remove when the LearningUnitYear.learning_container_year_id will be null=false
        if learning_unit_yr.learning_container_year else "",
        xls_build.translate(learning_unit_yr.subtype),
        learning_unit_yr.allocation_entity,
        learning_unit_yr.requirement_entity,
        learning_unit_yr.credits, xls_build.translate(learning_unit_yr.status)
    ]


def create_xls(user, found_learning_units, filters):
    titles = learning_unit_titles_part_1(display_proposal=True) + learning_unit_titles_part2()
    working_sheets_data = prepare_ue_xls_content(found_learning_units)
    parameters = {xls_build.DESCRIPTION: XLS_DESCRIPTION,
                  xls_build.USER: get_name_or_username(user),
                  xls_build.FILENAME: XLS_FILENAME,
                  xls_build.HEADER_TITLES: titles,
                  xls_build.WS_TITLE: WORKSHEET_TITLE}

    return xls_build.generate_xls(xls_build.prepare_xls_parameters_list(working_sheets_data, parameters), filters)


def create_xls_attributions(user, found_learning_units, filters):
    titles = learning_unit_titles_part1() + learning_unit_titles_part2() + [str(_('Tutor')),
                                                                            "{} ({})".format(str(_('Tutor')),
                                                                                             str(_('email'))),
                                                                            str(_('Function')),
                                                                            str(_('Substitute')),
                                                                            str(_('Beg. of attribution')),
                                                                            str(_('Attribution duration')),
                                                                            str(_('Attrib. vol1')),
                                                                            str(_('Attrib. vol2')),
                                                                            ]
    xls_data = prepare_xls_content_with_attributions(found_learning_units, len(titles))
    working_sheets_data = xls_data.get('data')
    cells_with_top_border = xls_data.get('cells_with_top_border')
    cells_with_white_font = xls_data.get('cells_with_white_font')
    parameters = {xls_build.DESCRIPTION: _('Learning units list with attributions'),
                  xls_build.USER: get_name_or_username(user),
                  xls_build.FILENAME: XLS_FILENAME,
                  xls_build.HEADER_TITLES: titles,
                  xls_build.WS_TITLE: WORKSHEET_TITLE,
                  xls_build.BORDER_CELLS: {xls_build.BORDER_TOP: cells_with_top_border},
                  xls_build.FONT_CELLS: {Font(color=Color('00FFFFFF')): cells_with_white_font},
                  xls_build.FONT_ROWS: {BOLD_FONT: [0]}
                  }

    return xls_build.generate_xls(xls_build.prepare_xls_parameters_list(working_sheets_data, parameters), filters)


def prepare_xls_content_with_attributions(found_learning_units: QuerySet, nb_columns: int) -> Dict:
    data = []
    qs = annotate_qs(found_learning_units)
    cells_with_top_border = []
    cells_with_white_font = []
    line = 2

    for learning_unit_yr in qs:
        first = True
        cells_with_top_border.extend(["{}{}".format(letter, line) for letter in _get_all_columns_reference(nb_columns)])

        lu_data_part1 = get_data_part1(learning_unit_yr, is_external_ue_list=False)
        lu_data_part2 = get_data_part2(learning_unit_yr, with_attributions=False)

        lu_data_part1.extend(lu_data_part2)

        attributions_values = attribution_charge_new.find_attribution_charge_new_by_learning_unit_year_as_dict(
            learning_unit_yr).values()
        if attributions_values:
            for value in attributions_values:
                data.append(lu_data_part1+_get_attribution_detail(value))
                line += 1
                if not first:
                    cells_with_white_font.extend(
                        ["{}{}".format(letter, line-1) for letter in _get_all_columns_reference(24)]
                    )
                first = False
        else:
            data.append(lu_data_part1)
            line += 1

    return {
        'data': data,
        'cells_with_top_border': cells_with_top_border or None,
        'cells_with_white_font': cells_with_white_font or None,
    }


def _get_attribution_detail(an_attribution):
    return [
        an_attribution.get('person').full_name,
        an_attribution.get('person').email,
        Functions[an_attribution['function']].value if 'function' in an_attribution else '',
        an_attribution.get('substitute') if an_attribution.get('substitute') else '',
        an_attribution.get('start_year'),
        an_attribution.get('duration') if an_attribution.get('duration') else '',
        an_attribution.get('LECTURING'),
        an_attribution.get('PRACTICAL_EXERCISES')
    ]


def volume_information(learning_unit_yr):
    return [
        get_significant_volume(learning_unit_yr.pm_vol_tot or 0),
        get_significant_volume(learning_unit_yr.pm_vol_q1 or 0),
        get_significant_volume(learning_unit_yr.pm_vol_q2 or 0),
        learning_unit_yr.pm_classes or 0,
        get_significant_volume(learning_unit_yr.pp_vol_tot or 0),
        get_significant_volume(learning_unit_yr.pp_vol_q1 or 0),
        get_significant_volume(learning_unit_yr.pp_vol_q2 or 0),
        learning_unit_yr.pp_classes or 0
    ]


# FIXME :: à discuter de la manière de faire à cause de code presque dupliqué
def acronym_with_version_label(acronym: str, transition_name: str, version_name: str) -> str:
    identity = ProgramTreeVersionIdentity(
        offer_acronym=acronym, transition_name=transition_name, version_name=version_name, year=None
    )
    version_str = version_label(identity)
    return "{}{}".format(acronym, version_str)


def _external_ue_titles() -> List[str]:
    return [
        str(_('Country')),
        str(_('City of institution')),
        str(_('Reference institution')),
        str(_('External code')),
        str(_('Url')),
        str(_('Local credits')),
    ]


def _get_external_ue_data(learning_unit_yr: LearningUnitYear) -> List['str']:
    organization = learning_unit_yr.campus.organization
    external_learning_unit_yr = learning_unit_yr.externallearningunityear
    return [
        organization.country or '',
        organization.main_address.city if organization.main_address else '',
        organization.name,
        external_learning_unit_yr.external_acronym,
        external_learning_unit_yr.url or '',
        "{0:.2f}".format(external_learning_unit_yr.external_credits)
        if external_learning_unit_yr.external_credits else ''
    ]


def _prepare_titles(is_external_ue_list: bool, with_attributions: bool, with_grp: bool) -> List[str]:
    titles = learning_unit_titles_part_1(not is_external_ue_list)
    titles_part2 = learning_unit_titles_part2()
    if with_grp:
        titles_part2.append(str(HEADER_PROGRAMS))
    if with_attributions:
        titles.extend(HEADER_TEACHERS)
    titles.extend(titles_part2)
    if is_external_ue_list:
        titles.extend(_external_ue_titles())
    return titles


def _build_complete_name(person: Person) -> str:
    return " ".join([
        ("{}".format(person.last_name) if person.last_name else "").upper(),
        person.first_name or ""
    ]).strip()


def _get_teachers(learning_unit_yr: LearningUnitYear) -> List[Person]:
    attributions = attribution_charge_new.find_attribution_charge_new_by_learning_unit_year_as_dict(
        learning_unit_yr)
    teachers = set()
    for k, attribution in attributions.items():
        if attribution.get('person'):
            teachers.add(attribution.get('person'))
    return teachers
