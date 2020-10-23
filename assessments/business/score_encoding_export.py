##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2020 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
from django.http import HttpResponse
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, colors
from openpyxl.writer.excel import save_virtual_workbook

from assessments.business.enrollment_state import get_line_color, ENROLLED_LATE_COLOR, NOT_ENROLLED_COLOR
from base import models as mdl
from base.models.enums import exam_enrollment_justification_type
from openpyxl.styles.borders import Border, Side, BORDER_MEDIUM
from openpyxl.styles import Style
from base.models.enums import peps_type
from base.models.student_specific_profile import StudentSpecificProfile
from osis_common.decorators.download import set_download_cookie

HEADER = [_('Academic year'), _('Session'), _('Learning unit'), _('Program'), _('Registration number'), _('Lastname'),
          _('Firstname'), _('Email'), _('Numbered scores'), _('Justification (A,T)'), _('End date Prof'),
          _('Type of specific profile'), _('Extra time (33% generally)'), _('Large print'),
          _('Specific room of examination'), _('Other educational facilities'), _('Educational tutor'),
          ]

JUSTIFICATION_ALIASES = {
    exam_enrollment_justification_type.ABSENCE_JUSTIFIED: "M",
    exam_enrollment_justification_type.ABSENCE_UNJUSTIFIED: "S",
    exam_enrollment_justification_type.CHEATING: "T",
}

FIRST_COL_LEGEND_ENROLLMENT_STATUS = 7
FIRST_ROW_LEGEND_ENROLLMENT_STATUS = 7

FIRST_COL_PEPS = 'L'
FIRST_ROW_PEPS = 12
LAST_COLUMN_NUMBER = 18

BORDER_LEFT = Border(
        left=Side(border_style=BORDER_MEDIUM,
                  color=Color('FF000000'),
                  ),
    )


@set_download_cookie
def export_xls(exam_enrollments):
    workbook = Workbook()
    worksheet = workbook.active
    _add_header_and_legend_to_file(exam_enrollments, worksheet)

    row_number = 12
    for exam_enroll in exam_enrollments:
        student = exam_enroll.learning_unit_enrollment.student
        offer = exam_enroll.learning_unit_enrollment.offer
        person = mdl.person.find_by_id(student.person.id)
        end_date = __get_session_exam_deadline(exam_enroll)

        score = None
        if exam_enroll.score_final is not None:
            if exam_enroll.session_exam.learning_unit_year.decimal_scores:
                score = "{0:.2f}".format(exam_enroll.score_final)
            else:
                score = "{0:.0f}".format(exam_enroll.score_final)

        justification = JUSTIFICATION_ALIASES.get(exam_enroll.justification_final, "")
        student_specific_profile = None
        if hasattr(exam_enroll.learning_unit_enrollment.student, 'studentspecificprofile'):
            student_specific_profile = exam_enroll.learning_unit_enrollment.student.studentspecificprofile

        line_content = [
            str(exam_enroll.learning_unit_enrollment.learning_unit_year.academic_year),
            str(exam_enroll.session_exam.number_session),
            exam_enroll.session_exam.learning_unit_year.acronym,
            offer.acronym,
            student.registration_id,
            person.last_name,
            person.first_name,
            person.email,
            score,
            str(justification),
            end_date if exam_enroll.enrollment_state == 'ENROLLED' else ''
        ]
        if student_specific_profile:
            line_content.extend([
                _get_type_peps(student_specific_profile),
                str(_('Yes')) if student_specific_profile.arrangement_additional_time else '-',
                str(_('Yes')) if student_specific_profile.arrangement_appropriate_copy else '-',
                str(_('Yes')) if student_specific_profile.arrangement_specific_locale else '-',
                str(_('Yes')) if student_specific_profile.arrangement_other else '-',
                str(student_specific_profile.guide) if student_specific_profile.guide else '-',
                ])
        else:
            line_content.extend(["-", "-", "-", "-", "-", "-"])
        worksheet.append(line_content)
        row_number += 1
        __coloring_non_editable(worksheet, row_number, score, exam_enroll.justification_final)
        _coloring_enrollment_state(worksheet, row_number, exam_enroll)
        _set_peps_border(worksheet, row_number)

    lst_exam_enrollments = list(exam_enrollments)
    number_session = lst_exam_enrollments[0].session_exam.number_session
    learn_unit_acronym = lst_exam_enrollments[0].session_exam.learning_unit_year.acronym
    academic_year = lst_exam_enrollments[0].learning_unit_enrollment.learning_unit_year.academic_year

    filename = "session_%s_%s_%s.xlsx" % (str(academic_year.year), str(number_session), learn_unit_acronym)
    response = HttpResponse(save_virtual_workbook(workbook),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    return response


def _add_header_and_legend_to_file(exam_enrollments, worksheet):
    ue = exam_enrollments[0].learning_unit_enrollment.learning_unit_year
    worksheet.append([str(ue) + " " + ue.complete_title if ue.complete_title else str(ue)])
    worksheet.append([str('Session: %s' % exam_enrollments[0].session_exam.number_session)])
    worksheet.append([str('')])
    __display_creation_date_with_message_about_state(worksheet, row_number=4)
    __display_warning_about_students_deliberated(worksheet, row_number=5)
    worksheet.append([str('')])
    show_decimal = ue.decimal_scores
    __display_legends(worksheet, show_decimal)
    _color_legend(worksheet)
    worksheet.append([str('')])
    __columns_resizing(worksheet)
    header_translate_list = [str(elem) for elem in HEADER]
    worksheet.append(header_translate_list)


def __columns_resizing(ws):
    """
    Definition of the columns sizes
    """
    col_academic_year = ws.column_dimensions['A']
    col_academic_year.width = 18
    col_academic_year = ws.column_dimensions['C']
    col_academic_year.width = 18
    col_academic_year = ws.column_dimensions['E']
    col_academic_year.width = 18
    col_last_name = ws.column_dimensions['F']
    col_last_name.width = 25
    col_first_name = ws.column_dimensions['G']
    col_first_name.width = 25
    col_email = ws.column_dimensions['H']
    col_email.width = 30
    col_note = ws.column_dimensions['I']
    col_note.width = 15
    col_note = ws.column_dimensions['J']
    col_note.width = 15
    col_note = ws.column_dimensions['K']
    col_note.width = 15
    col_note = ws.column_dimensions['L']
    col_note.width = 20
    col_note = ws.column_dimensions['M']
    col_note.width = 25
    col_note = ws.column_dimensions['N']
    col_note.width = 15
    col_note = ws.column_dimensions['O']
    col_note.width = 25
    col_note = ws.column_dimensions['P']
    col_note.width = 25
    col_note = ws.column_dimensions['Q']
    col_note.width = 30


def __coloring_non_editable(ws, row_number, score, justification):
    """
    Coloring of the non-editable columns
    """
    pattern_fill_grey = PatternFill(patternType='solid', fgColor=Color('C1C1C1'))
    column_number = 1
    while column_number < LAST_COLUMN_NUMBER:
        if column_number < 9 or column_number > 10:
            ws.cell(row=row_number, column=column_number).fill = pattern_fill_grey
        else:
            if not(score is None and justification is None):
                ws.cell(row=row_number, column=9).fill = pattern_fill_grey
                ws.cell(row=row_number, column=10).fill = pattern_fill_grey

        column_number += 1


def __display_creation_date_with_message_about_state(ws, row_number):
    date_format = str(_('date_format'))
    printing_date = timezone.now()
    printing_date = printing_date.strftime(date_format)

    ws.cell(row=row_number, column=1).value = str(
        '%s' % (_("The data presented on this document correspond to the state of the system dated %(printing_date)s "
                  "and are likely to evolve") % {'printing_date': printing_date}))
    ws.cell(row=row_number, column=1).font = Font(color=colors.RED)


def __display_warning_about_students_deliberated(ws, row_number):
    ws.cell(row=row_number, column=1).value = str(_("Students deliberated are not shown"))
    ws.cell(row=row_number, column=1).font = Font(color=colors.RED)


def __display_legends(ws, decimal):
    ws.append([
        str(_('Justification')),
        str(_("Accepted value: %(justification_label_authorized)s ")
            % {"justification_label_authorized": mdl.exam_enrollment.justification_label_authorized()}),
        str(''),
        str(''),
        str(''),
        str(''),
        str(_('Enrolled lately')),
    ])
    ws.append([
        str(''),
        str(_("Other values reserved to administration: %(justification_other_values)s ") % {
            'justification_other_values': justification_other_values()}),
        str(''),
        str(''),
        str(''),
        str(''),
        str(_('Unsubscribed lately')),
    ])
    ws.append([
        str(_('Numbered scores')),
        str(_('Score legend: %(score_legend)s (0=Score of presence)') % {"score_legend": "0 - 20"}),
    ])
    ws.append([
            str(''),
            str(_('Decimals authorized for this learning unit'))
            if decimal else
            str(_('Unauthorized decimal for this learning unit'))
    ])


def justification_other_values():
    return "%s, %s" % (_('S=Unjustified Absence'),
                       _('M=Justified Absence'))


def __get_session_exam_deadline(exam_enroll):
    date_format = str(_('date_format'))
    deadline = None

    session_exam_deadline = mdl.exam_enrollment.get_session_exam_deadline(exam_enroll)
    if session_exam_deadline:
        deadline = session_exam_deadline.deadline_tutor_computed if session_exam_deadline.deadline_tutor_computed else\
                   session_exam_deadline.deadline

    return deadline.strftime(date_format) if deadline else "-"


def _coloring_enrollment_state(ws, row_number, exam_enroll):
    """
    Coloring of the non-editable columns
    """
    enrollment_state_color = get_line_color(exam_enroll)
    if enrollment_state_color:
        pattern_fill_enrollment_state = PatternFill(patternType='solid',
                                                    fgColor=enrollment_state_color.lstrip("#"))
        column_number = 1
        while column_number < LAST_COLUMN_NUMBER:
            ws.cell(row=row_number, column=column_number).fill = pattern_fill_enrollment_state
            column_number += 1


def _color_legend(ws):
    __apply_style_to_cells(ws, ENROLLED_LATE_COLOR, FIRST_ROW_LEGEND_ENROLLMENT_STATUS)
    __apply_style_to_cells(ws, NOT_ENROLLED_COLOR, FIRST_ROW_LEGEND_ENROLLMENT_STATUS + 1)
    ws.cell(row=10, column=2).font = Font(color=colors.RED)


def __apply_style_to_cells(ws, color_style, row):
    fill_pattern = PatternFill(
        patternType='solid',
        fgColor=Color(color_style.lstrip("#"))
    )
    ws.cell(row=row, column=FIRST_COL_LEGEND_ENROLLMENT_STATUS).fill = fill_pattern


def _set_peps_border(ws, last_row_number):
    """
    Set border at the left of the first peps column
    """
    cpt = FIRST_ROW_PEPS
    while cpt <= last_row_number:
        cell = ws["{}{}".format(FIRST_COL_PEPS, cpt)]
        _update_border_for_first_peps_column(cell)
        cpt += 1


def _get_type_peps(student_specific_profile: StudentSpecificProfile) -> str:
    if student_specific_profile.type == peps_type.PepsTypes.SPORT.name:
        return "{} - {}".format(
            str(_(student_specific_profile.get_type_display())) or "-",
            str(_(student_specific_profile.get_subtype_sport_display())) or "-",
        )
    if student_specific_profile.type == peps_type.PepsTypes.DISABILITY.name:
        return "{} - {}".format(
            str(_(student_specific_profile.get_type_display())) or "-",
            str(_(student_specific_profile.get_subtype_disability_display())) or "-",
        )

    return str(_(student_specific_profile.get_type_display())) or "-"


def _update_border_for_first_peps_column(cell):
    if cell.has_style:
        c = cell.style
        c.border = BORDER_LEFT
        cell.style = c
