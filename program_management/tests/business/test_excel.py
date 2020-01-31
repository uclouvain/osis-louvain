##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2019 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
import html
from unittest import mock

from django.test import TestCase
from django.utils.translation import gettext_lazy as _
from openpyxl.styles import Style, Font

from attribution.tests.factories.attribution_charge_new import AttributionChargeNewFactory
from attribution.tests.factories.attribution_new import AttributionNewFactory
from base.business.learning_unit_xls import CREATION_COLOR, MODIFICATION_COLOR, TRANSFORMATION_COLOR, \
    TRANSFORMATION_AND_MODIFICATION_COLOR, SUPPRESSION_COLOR
from base.models.enums.prerequisite_operator import AND, OR
from base.tests.factories.business.learning_units import GenerateContainer
from base.tests.factories.education_group_year import EducationGroupYearFactory
from base.tests.factories.group_element_year import GroupElementYearChildLeafFactory, GroupElementYearFactory
from base.tests.factories.learning_achievement import LearningAchievementFactory
from base.tests.factories.learning_component_year import LecturingLearningComponentYearFactory, \
    PracticalLearningComponentYearFactory
from base.tests.factories.person import PersonFactory
from base.tests.factories.prerequisite import PrerequisiteFactory
from base.tests.factories.proposal_learning_unit import ProposalLearningUnitFactory
from base.tests.factories.teaching_material import TeachingMaterialFactory
from base.tests.factories.tutor import TutorFactory
from program_management.business.excel import EducationGroupYearLearningUnitsContainedToExcel
from program_management.business.excel import EducationGroupYearLearningUnitsPrerequisitesToExcel, \
    EducationGroupYearLearningUnitsIsPrerequisiteOfToExcel, _get_blocks_prerequisite_of, FIX_TITLES, _get_headers, \
    optional_header_for_proposition, optional_header_for_credits, optional_header_for_volume, _get_attribution_line, \
    optional_header_for_required_entity, optional_header_for_active, optional_header_for_allocation_entity, \
    optional_header_for_description_fiche, optional_header_for_english_title, optional_header_for_language, \
    optional_header_for_periodicity, optional_header_for_quadrimester, optional_header_for_session_derogation, \
    optional_header_for_specifications, optional_header_for_teacher_list, \
    _fix_data, _get_workbook_for_custom_xls, _build_legend_sheet, LEGEND_WB_CONTENT, LEGEND_WB_STYLE, _optional_data, \
    _build_excel_lines_ues, _get_optional_data, BOLD_FONT, _build_specifications_cols, _build_description_fiche_cols, \
    _build_validate_html_list_to_string
from program_management.business.utils import html2text
from program_management.forms.custom_xls import CustomXlsForm
from reference.tests.factories.language import LanguageFactory

CMS_TXT_WITH_LIST = '<ol> ' \
                    '<li>La structure atomique de la mati&egrave;re</li> ' \
                    '<li>Les diff&eacute;rentes structures mol&eacute;culaires</li> ' \
                    '</ol>'
CMS_TXT_WITH_LIST_AFTER_FORMATTING = 'La structure atomique de la matière\n' \
                                    'Les différentes structures moléculaires'

CMS_TXT_WITH_LINK = '<a href="https://moodleucl.uclouvain.be">moodle</a>'
CMS_TXT_WITH_LINK_AFTER_FORMATTING = 'moodle - [https://moodleucl.uclouvain.be] \n'



class TestGeneratePrerequisitesWorkbook(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.education_group_year = EducationGroupYearFactory()
        cls.child_leaves = GroupElementYearChildLeafFactory.create_batch(
            6,
            parent=cls.education_group_year
        )
        luy_acronyms = ["LCORS124" + str(i) for i in range(0, len(cls.child_leaves))]
        for node, acronym in zip(cls.child_leaves, luy_acronyms):
            node.child_leaf.acronym = acronym
            node.child_leaf.save()

        cls.luy_children = [child.child_leaf for child in cls.child_leaves]

        PrerequisiteFactory(
            learning_unit_year=cls.luy_children[0],
            education_group_year=cls.education_group_year,
            items__groups=(
                (cls.luy_children[1],),
            )
        )
        PrerequisiteFactory(
            learning_unit_year=cls.luy_children[2],
            education_group_year=cls.education_group_year,
            items__groups=(
                (cls.luy_children[3],),
                (cls.luy_children[4], cls.luy_children[5])
            )
        )
        cls.workbook_prerequisites = \
            EducationGroupYearLearningUnitsPrerequisitesToExcel(cls.education_group_year)._to_workbook()
        cls.workbook_is_prerequisite = \
            EducationGroupYearLearningUnitsIsPrerequisiteOfToExcel(cls.education_group_year)._to_workbook()
        cls.sheet_prerequisites = cls.workbook_prerequisites.worksheets[0]
        cls.sheet_is_prerequisite = cls.workbook_is_prerequisite.worksheets[0]

    def test_header_lines(self):
        expected_headers = [
            [self.education_group_year.acronym, self.education_group_year.title, _('Code'), _('Title'),
             _('Cred. rel./abs.'), _('Block'), _('Mandatory')],
            [_("Official"), None, None, None, None, None, None]
        ]

        headers = [row_to_value(row) for row in self.sheet_prerequisites.iter_rows(range_string="A1:G2")]
        self.assertListEqual(headers, expected_headers)

    def test_when_learning_unit_year_has_one_prerequisite(self):
        expected_content = [
            [self.luy_children[0].acronym, self.luy_children[0].complete_title, None, None, None, None, None],

            [_("has as prerequisite") + " :", '',
             self.luy_children[1].acronym,
             self.luy_children[1].complete_title_i18n,
             "{} / {}".format(self.child_leaves[1].relative_credits, self.luy_children[1].credits),
             str(self.child_leaves[1].block) if self.child_leaves[1].block else '',
             _("Yes") if self.child_leaves[1].is_mandatory else _("No")]
        ]

        content = [row_to_value(row) for row in self.sheet_prerequisites.iter_rows(range_string="A3:G4")]
        self.assertListEqual(expected_content, content)

    def test_when_learning_unit_year_has_multiple_prerequisites(self):
        expected_content = [
            [self.luy_children[2].acronym, self.luy_children[2].complete_title, None, None, None, None, None],

            [_("has as prerequisite") + " :", '', self.luy_children[3].acronym,
             self.luy_children[3].complete_title_i18n,
             "{} / {}".format(self.child_leaves[3].relative_credits, self.luy_children[3].credits),
             str(self.child_leaves[3].block) if self.child_leaves[3].block else '',
             _("Yes") if self.child_leaves[3].is_mandatory else _("No")],

            ['', _(AND), "(" + self.luy_children[4].acronym, self.luy_children[4].complete_title_i18n,
             "{} / {}".format(self.child_leaves[4].relative_credits, self.luy_children[4].credits),
             str(self.child_leaves[4].block) if self.child_leaves[4].block else '',
             _("Yes") if self.child_leaves[4].is_mandatory else _("No")
             ],

            ['', _(OR), self.luy_children[5].acronym + ")", self.luy_children[5].complete_title_i18n,
             "{} / {}".format(self.child_leaves[5].relative_credits, self.luy_children[5].credits),
             str(self.child_leaves[5].block) if self.child_leaves[5].block else '',
             _("Yes") if self.child_leaves[5].is_mandatory else _("No")
             ]
        ]
        content = [row_to_value(row) for row in self.sheet_prerequisites.iter_rows(range_string="A5:G8")]
        self.assertListEqual(expected_content, content)

    def test_get_blocks_prerequisite_of(self):
        gey = GroupElementYearFactory(block=123)
        self.assertEqual(_get_blocks_prerequisite_of(gey), '1 ; 2 ; 3')
        gey = GroupElementYearFactory(block=1)
        self.assertEqual(_get_blocks_prerequisite_of(gey), '1')


def row_to_value(sheet_row):
    return [cell.value for cell in sheet_row]


class TestGenerateEducationGroupYearLearningUnitsContainedWorkbook(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.education_group_year = EducationGroupYearFactory()
        cls.child_leaves = GroupElementYearChildLeafFactory.create_batch(
            2,
            parent=cls.education_group_year,
            is_mandatory=True
        )
        for node, acronym in zip(cls.child_leaves, ["LCORS124" + str(i) for i in range(0, len(cls.child_leaves))]):
            node.child_leaf.acronym = acronym
            node.child_leaf.save()

        cls.luy_children = [child.child_leaf for child in cls.child_leaves]
        cls.workbook_contains = \
            EducationGroupYearLearningUnitsContainedToExcel(cls.education_group_year, CustomXlsForm({}))._to_workbook()
        cls.sheet_contains = cls.workbook_contains.worksheets[0]

        generator_container = GenerateContainer(cls.education_group_year.academic_year,
                                                cls.education_group_year.academic_year)
        cls.luy = generator_container.generated_container_years[0].learning_unit_year_full

        cls.lecturing_component = LecturingLearningComponentYearFactory(
            learning_unit_year=cls.luy)
        cls.practical_component = PracticalLearningComponentYearFactory(
            learning_unit_year=cls.luy)
        cls.person_1 = PersonFactory(last_name='Dupont', first_name="Marcel", email="dm@gmail.com")
        cls.person_2 = PersonFactory(last_name='Marseillais', first_name="Pol", email="pm@gmail.com")
        cls.tutor_1 = TutorFactory(person=cls.person_1)
        cls.tutor_2 = TutorFactory(person=cls.person_2)
        cls.attribution_1 = AttributionNewFactory(
            tutor=cls.tutor_1,
            learning_container_year=cls.luy.learning_container_year
        )
        cls.charge_lecturing = AttributionChargeNewFactory(
            attribution=cls.attribution_1,
            learning_component_year=cls.lecturing_component
        )
        cls.charge_practical = AttributionChargeNewFactory(
            attribution=cls.attribution_1,
            learning_component_year=cls.practical_component
        )
        cls.attribution_2 = AttributionNewFactory(
            tutor=cls.tutor_2,
            learning_container_year=cls.luy.learning_container_year
        )
        cls.charge_lecturing = AttributionChargeNewFactory(
            attribution=cls.attribution_2,
            learning_component_year=cls.lecturing_component
        )
        cls.charge_practical = AttributionChargeNewFactory(
            attribution=cls.attribution_2,
            learning_component_year=cls.practical_component
        )
        cls.gey = GroupElementYearChildLeafFactory(
            child_leaf=cls.luy
        )

    def test_header_lines_without_optional_titles(self):
        custom_xls_form = CustomXlsForm({})
        expected_headers = FIX_TITLES

        self.assertListEqual(_get_headers(custom_xls_form)[0], expected_headers)

    def test_header_lines_with_optional_titles(self):
        custom_xls_form = CustomXlsForm({
            'required_entity': 'on',
            'allocation_entity': 'on',
            'credits': 'on',
            'periodicity': 'on',
            'active': 'on',
            'quadrimester': 'on',
            'session_derogation': 'on',
            'volume': 'on',
            'teacher_list': 'on',
            'proposition': 'on',
            'english_title': 'on',
            'language': 'on',
            'specifications': 'on',
            'description_fiche': 'on',
        }
        )

        expected_headers = \
            FIX_TITLES + optional_header_for_required_entity + optional_header_for_allocation_entity +  \
            optional_header_for_credits + optional_header_for_periodicity + optional_header_for_active + \
            optional_header_for_quadrimester + optional_header_for_session_derogation + optional_header_for_volume + \
            optional_header_for_teacher_list + optional_header_for_proposition + optional_header_for_english_title + \
            optional_header_for_language + optional_header_for_specifications + optional_header_for_description_fiche
        self.assertListEqual(_get_headers(custom_xls_form)[0], expected_headers)

    def test_get_attribution_line(self):
        person = PersonFactory(last_name='Last', first_name='First', middle_name='Middle')
        self.assertEqual(_get_attribution_line(person), 'LAST First Middle')
        person = PersonFactory(last_name=None, first_name='First', middle_name='Middle')
        self.assertEqual(_get_attribution_line(person), 'First Middle')
        self.assertEqual(_get_attribution_line(None), '')

    def test_fix_data(self):
        gey = self.child_leaves[0]
        luy = self.luy_children[0]
        expected = get_expected_data(gey, luy)
        res = _fix_data(gey, luy)
        self.assertEqual(res, expected)

    def test_legend_workbook_exists(self):
        wb = _get_workbook_for_custom_xls([['header'], [['row1 col1']]], True, {})
        self.assertEqual(len(wb.worksheets), 2)

    def test_legend_workbook_do_not_exists(self):
        wb = _get_workbook_for_custom_xls([['header'], [['row1 col1']]], False, {})
        self.assertEqual(len(wb.worksheets), 1)

    def test_legend_workbook_content(self):
        expected_content = [[_("Creation")], [_("Modification")], [_("Transformation")],
                            [_("Transformation and modification")], [_("Suppression")]]

        data = _build_legend_sheet()
        self.assertListEqual(data.get(LEGEND_WB_CONTENT), expected_content)

    def test_legend_workbook_style(self):
        data = _build_legend_sheet()
        self.assertListEqual(data.get(LEGEND_WB_STYLE).get(Style(font=Font(color=CREATION_COLOR))), [1])
        self.assertListEqual(data.get(LEGEND_WB_STYLE).get(Style(font=Font(color=MODIFICATION_COLOR))), [2])
        self.assertListEqual(data.get(LEGEND_WB_STYLE).get(Style(font=Font(color=TRANSFORMATION_COLOR))), [3])
        self.assertListEqual(
            data.get(LEGEND_WB_STYLE).get(Style(font=Font(color=TRANSFORMATION_AND_MODIFICATION_COLOR))),
            [4])
        self.assertListEqual(data.get(LEGEND_WB_STYLE).get(Style(font=Font(color=SUPPRESSION_COLOR))), [5])

    def test_no_optional_data_to_add(self):
        form = CustomXlsForm({})
        self.assertDictEqual(_optional_data(form),
                             {'has_required_entity': False,
                              'has_proposition': False,
                              'has_credits': False,
                              'has_allocation_entity': False,
                              'has_english_title': False,
                              'has_teacher_list': False,
                              'has_periodicity': False,
                              'has_active': False,
                              'has_volume': False,
                              'has_quadrimester': False,
                              'has_session_derogation': False,
                              'has_language': False,
                              'has_description_fiche': False,
                              'has_specifications': False,
                              }
                             )

    def test_all_optional_data_to_add(self):
        form = CustomXlsForm({'required_entity': 'on',
                              'proposition': 'on',
                              'credits': 'on',
                              'allocation_entity': 'on',
                              'english_title': 'on',
                              'teacher_list': 'on',
                              'periodicity': 'on',
                              'active': 'on',
                              'volume': 'on',
                              'quadrimester': 'on',
                              'session_derogation': 'on',
                              'language': 'on',
                              'description_fiche': 'on',
                              'specifications': 'on',
                              })
        self.assertDictEqual(_optional_data(form),
                             {'has_required_entity': True,
                              'has_proposition': True,
                              'has_credits': True,
                              'has_allocation_entity': True,
                              'has_english_title': True,
                              'has_teacher_list': True,
                              'has_periodicity': True,
                              'has_active': True,
                              'has_volume': True,
                              'has_quadrimester': True,
                              'has_session_derogation': True,
                              'has_language': True,
                              'has_description_fiche': True,
                              'has_specifications': True,
                              }
                             )

    def test_data(self):
        custom_form = CustomXlsForm({})
        exl = EducationGroupYearLearningUnitsContainedToExcel(self.education_group_year, custom_form)
        data = _build_excel_lines_ues(custom_form, exl.learning_unit_years_parent)
        content = data.get('content')
        self._assert_content_equals(content, exl)
        # First line (Header line) is always bold
        self.assertListEqual(data.get('colored_cells')[Style(font=BOLD_FONT)], [0])

    def _assert_content_equals(self, content, exl):
        idx = 1
        for gey in exl.learning_unit_years_parent:
            luy = gey.child_leaf
            expected = get_expected_data(gey, luy)
            self.assertListEqual(content[idx], expected)
            idx += 1

    def test_get_optional_required_entity(self):
        optional_data = initialize_optional_data()
        optional_data['has_required_entity'] = True
        self.assertCountEqual(_get_optional_data([], self.luy, optional_data, self.gey),
                              [self.luy.learning_container_year.requirement_entity])

    def test_get_optional_allocation_entity(self):
        optional_data = initialize_optional_data()
        optional_data['has_allocation_entity'] = True
        self.assertCountEqual(_get_optional_data([], self.luy, optional_data, self.gey),
                              [self.luy.learning_container_year.allocation_entity])

    def test_get_optional_credits(self):
        optional_data = initialize_optional_data()
        optional_data['has_credits'] = True

        self.assertCountEqual(_get_optional_data([], self.luy, optional_data, self.gey),
                              [self.luy.credits.to_integral_value()])

    def test_get_optional_has_periodicity(self):
        optional_data = initialize_optional_data()
        optional_data['has_periodicity'] = True
        self.assertCountEqual(_get_optional_data([], self.luy, optional_data, self.gey),
                              [self.luy.get_periodicity_display()])

    def test_get_optional_has_active(self):
        optional_data = initialize_optional_data()
        optional_data['has_active'] = True
        self.assertCountEqual(_get_optional_data([], self.luy, optional_data, self.gey),
                              [_('yes')])

    def test_get_optional_has_quadrimester(self):
        optional_data = initialize_optional_data()
        optional_data['has_quadrimester'] = True
        self.assertCountEqual(_get_optional_data([], self.luy, optional_data, self.gey),
                              [self.luy.get_quadrimester_display() or ''])

    def test_get_optional_has_session_derogation(self):
        optional_data = initialize_optional_data()
        optional_data['has_session_derogation'] = True
        self.assertCountEqual(_get_optional_data([], self.luy, optional_data, self.gey),
                              [self.luy.get_session_display() or ''])

    def test_get_optional_has_proposition(self):
        optional_data = initialize_optional_data()
        optional_data['has_proposition'] = True
        self.assertCountEqual(_get_optional_data([], self.luy, optional_data, self.gey),
                              ['', ''])
        proposal = ProposalLearningUnitFactory(learning_unit_year=self.luy)

        self.assertCountEqual(_get_optional_data([], self.luy, optional_data, self.gey),
                              [proposal.get_type_display(), proposal.get_state_display()])

    def test_get_optional_has_english_title(self):
        optional_data = initialize_optional_data()
        optional_data['has_english_title'] = True
        self.assertCountEqual(_get_optional_data([], self.luy, optional_data, self.gey),
                              [self.luy.complete_title_english])

    def test_get_optional_has_language(self):
        optional_data = initialize_optional_data()
        optional_data['has_language'] = True
        self.assertCountEqual(_get_optional_data([], self.luy, optional_data, self.gey),
                              [self.luy.language])

    def test_get_optional_has_teacher_list(self):
        optional_data = initialize_optional_data()
        optional_data['has_teacher_list'] = True
        teacher_data = _get_optional_data([], self.luy, optional_data, self.gey)
        self.assertEqual(teacher_data[0], "{} {};{} {}"
                         .format(self.person_1.last_name.upper(), self.person_1.first_name,
                                 self.person_2.last_name.upper(), self.person_2.first_name))
        self.assertEqual(teacher_data[1], "{};{}"
                         .format(self.person_1.email,
                                 self.person_2.email))

    @mock.patch("program_management.business.excel._annotate_with_description_fiche_specifications")
    def test_get_optional_has_description_fiche_annotate_called(self, mock):
        optional_data = initialize_optional_data()
        optional_data['has_description_fiche'] = True

        custom_form = CustomXlsForm({'description_fiche': 'on'})
        EducationGroupYearLearningUnitsContainedToExcel(self.education_group_year, custom_form)
        self.assertTrue(mock.called)

    @mock.patch("program_management.business.excel._annotate_with_description_fiche_specifications")
    def test_get_optional_has_specifications_annotate_called(self, mock):
        optional_data = initialize_optional_data()
        optional_data['has_specifications'] = True

        custom_form = CustomXlsForm({'specifications': 'on'})
        EducationGroupYearLearningUnitsContainedToExcel(self.education_group_year, custom_form)
        self.assertTrue(mock.called)

    def test_build_description_fiche_cols(self):

        teaching_material_1 = TeachingMaterialFactory(
            learning_unit_year=self.luy, title='Title mandatory', mandatory=True
        )
        teaching_material_2 = TeachingMaterialFactory(
            learning_unit_year=self.luy, title='Title non-mandatory', mandatory=False
        )

        _initialize_cms_data_description_fiche(self.gey)

        description_fiche = _build_description_fiche_cols(self.luy, self.gey)

        self.assertEqual(description_fiche.resume, "{}".format(CMS_TXT_WITH_LIST_AFTER_FORMATTING))
        self.assertEqual(description_fiche.resume_en, "{}".format(CMS_TXT_WITH_LIST_AFTER_FORMATTING))
        self.assertEqual(description_fiche.teaching_methods, "{}".format(CMS_TXT_WITH_LIST_AFTER_FORMATTING))
        self.assertEqual(description_fiche.teaching_methods_en, "{}".format(CMS_TXT_WITH_LIST_AFTER_FORMATTING))
        self.assertEqual(description_fiche.evaluation_methods, "{}".format(CMS_TXT_WITH_LIST_AFTER_FORMATTING))
        self.assertEqual(description_fiche.evaluation_methods_en, "{}".format(CMS_TXT_WITH_LIST_AFTER_FORMATTING))
        self.assertEqual(description_fiche.other_informations, "{}".format(CMS_TXT_WITH_LIST_AFTER_FORMATTING))
        self.assertEqual(description_fiche.other_informations_en, "{}".format(CMS_TXT_WITH_LIST_AFTER_FORMATTING))
        self.assertEqual(description_fiche.mobility, "{}".format(CMS_TXT_WITH_LIST_AFTER_FORMATTING))
        self.assertEqual(description_fiche.bibliography, "{}".format(CMS_TXT_WITH_LIST_AFTER_FORMATTING))

        self.assertEqual(description_fiche.online_resources, "{}".format(CMS_TXT_WITH_LINK_AFTER_FORMATTING))
        self.assertEqual(description_fiche.online_resources_en, "{}".format(CMS_TXT_WITH_LINK_AFTER_FORMATTING))

        self.assertEqual(description_fiche.teaching_materials,
                         "{} - {}\n{} - {}".format(_('Mandatory'),
                                                   teaching_material_1.title,
                                                   _('Non-mandatory'),
                                                   teaching_material_2.title))

    def test_build_specifications_cols(self):

        lang_fr = LanguageFactory(code='FR')
        lang_en = LanguageFactory(code='EN')

        achievement_1_fr = LearningAchievementFactory(learning_unit_year=self.luy, language=lang_fr)
        achievement_2_fr = LearningAchievementFactory(learning_unit_year=self.luy, language=lang_fr)
        achievement_1_en = LearningAchievementFactory(learning_unit_year=self.luy, language=lang_en)
        LearningAchievementFactory(learning_unit_year=self.luy, language=lang_en, text="    ")
        LearningAchievementFactory(learning_unit_year=self.luy, language=lang_en, text="    ", code_name=None)

        initialize_cms_specifications_data_description_fiche(self.gey)
        specifications_data = _build_specifications_cols(self.luy, self.gey)

        self.assertEqual(specifications_data.prerequisite, CMS_TXT_WITH_LIST_AFTER_FORMATTING)
        self.assertEqual(specifications_data.prerequisite_en, CMS_TXT_WITH_LIST_AFTER_FORMATTING)
        self.assertEqual(specifications_data.themes_discussed, CMS_TXT_WITH_LIST_AFTER_FORMATTING)
        self.assertEqual(specifications_data.themes_discussed_en, CMS_TXT_WITH_LIST_AFTER_FORMATTING)
        self.assertEqual(specifications_data.achievements_fr, "{} -{}\n{} -{}".format(
            achievement_1_fr.code_name, achievement_1_fr.text,
            achievement_2_fr.code_name, achievement_2_fr.text)
                         )
        self.assertEqual(specifications_data.achievements_en, "{} -{}".format(
            achievement_1_en.code_name, achievement_1_en.text)
                         )

    def test_build_validate_html_list_to_string(self):
        self.assertEqual(_build_validate_html_list_to_string(None, html2text), "")

    def test_build_validate_html_list_to_string_illegal_character(self):
        self.assertEqual(_build_validate_html_list_to_string("", html2text),
                         "!!! {}".format(_('IMPOSSIBLE TO DISPLAY BECAUSE OF AN ILLEGAL CHARACTER IN STRING')))

    def test_build_validate_html_list_to_string_wrong_method(self):
        self.assertEqual(_build_validate_html_list_to_string('Test', None), 'Test')
        self.assertEqual(_build_validate_html_list_to_string('Test', _get_blocks_prerequisite_of), 'Test')

    def test_row_height_not_populated(self):
        custom_form = CustomXlsForm({})
        exl = EducationGroupYearLearningUnitsContainedToExcel(self.education_group_year, custom_form)
        data = _build_excel_lines_ues(custom_form, exl.qs)
        self.assertDictEqual(data.get('row_height'), {})

    def test_row_height_populated(self):
        custom_form = CustomXlsForm({'description_fiche': 'on'})
        exl = EducationGroupYearLearningUnitsContainedToExcel(self.education_group_year, custom_form)
        data = _build_excel_lines_ues(custom_form, exl.qs)
        self.assertDictEqual(data.get('row_height'), {'height': 30, 'start': 2, 'stop': 4})

    def test_html_list_to_string(self):
        ch = '''<head></head>
                <body>
                <style type="text/css"></style>
                <ul>
                    <li>Cfr. Student corner</li>
                </ul>            
                <p>Cfr. Syllabus</p>
                <script>alert('coucou');</script>    
                </body>        
                '''
        expected = "Cfr. Student corner\nCfr. Syllabus"
        self.assertEqual(html2text(html.unescape(ch)), expected)

    def test_convert(self):
        res = html2text(html.unescape("<p>Introduire aux m&eacute;thodes d&#39;analyse</p>"))
        self.assertEqual(res, "Introduire aux méthodes d'analyse")

    def test_keep_UES_tree_order_in_qs(self):
        exl = EducationGroupYearLearningUnitsContainedToExcel(self.education_group_year, CustomXlsForm({}))
        expected_ids_following_tree_order = [lu.id for lu in exl.learning_unit_years_parent]
        ids_ordered_for_xls = [lu.id for lu in list(exl.qs)]
        self.assertCountEqual(expected_ids_following_tree_order, ids_ordered_for_xls)


def get_expected_data(gey, luy):
    expected = [luy.acronym,
                luy.academic_year,
                luy.complete_title_i18n,
                luy.get_container_type_display(),
                luy.get_subtype_display(),
                "{} - {}".format(gey.parent.partial_acronym, gey.parent.title),
                "{} / {}".format(gey.relative_credits or '-', luy.credits.to_integral_value() or '-'),
                gey.block or '',
                _('yes')

                ]
    return expected


def initialize_optional_data():
    return {
        'has_required_entity': False,
        'has_allocation_entity': False,
        'has_credits': False,
        'has_periodicity': False,
        'has_active': False,
        'has_quadrimester': False,
        'has_session_derogation': False,
        'has_volume': False,
        'has_teacher_list': False,
        'has_proposition': False,
        'has_english_title': False,
        'has_language': False,
        'has_description_fiche': False,
        'has_specifications': False,
    }


def _initialize_cms_data_description_fiche(gey):
    gey_cms = gey
    gey_cms.resume = CMS_TXT_WITH_LIST
    gey_cms.resume_en = CMS_TXT_WITH_LIST
    gey_cms.teaching_methods = CMS_TXT_WITH_LIST
    gey_cms.teaching_methods_en = CMS_TXT_WITH_LIST
    gey_cms.evaluation_methods = CMS_TXT_WITH_LIST
    gey_cms.evaluation_methods_en = CMS_TXT_WITH_LIST
    gey_cms.other_informations = CMS_TXT_WITH_LIST
    gey_cms.other_informations_en = CMS_TXT_WITH_LIST
    gey_cms.bibliography = CMS_TXT_WITH_LIST
    gey_cms.mobility = CMS_TXT_WITH_LIST

    gey_cms.online_resources = CMS_TXT_WITH_LINK
    gey_cms.online_resources_en = CMS_TXT_WITH_LINK

    return gey_cms


def initialize_cms_specifications_data_description_fiche(gey):
    gey_cms = gey
    gey_cms.prerequisite = CMS_TXT_WITH_LIST
    gey_cms.prerequisite_en = CMS_TXT_WITH_LIST
    gey_cms.themes_discussed = CMS_TXT_WITH_LIST
    gey_cms.themes_discussed_en = CMS_TXT_WITH_LIST
    return gey_cms
