##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2020 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
from collections import namedtuple, defaultdict
from typing import List

from django.utils import translation
from django.utils.translation import gettext as _
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Color, PatternFill, Font
from openpyxl.styles.borders import BORDER_THICK
from openpyxl.styles.colors import RED, GREEN
from openpyxl.writer.excel import save_virtual_workbook

from backoffice.settings.base import LANGUAGE_CODE_EN
from backoffice.settings.base import LEARNING_UNIT_PORTAL_URL
from base.models.enums.prerequisite_operator import OR, AND
from osis_common.document.xls_build import _build_worksheet, CONTENT_KEY, HEADER_TITLES_KEY, WORKSHEET_TITLE_KEY, \
    FILL_NO_GRAY, FONT_CELLS, FILL_CELLS
from program_management.ddd.business_types import *
from program_management.ddd.domain import link
from program_management.ddd.domain.node import NodeIdentity
from program_management.ddd.domain.prerequisite import Prerequisite, PrerequisiteItem
from program_management.ddd.domain.program_tree import ProgramTreeIdentity
from program_management.ddd.repositories.program_tree import ProgramTreeRepository

BORDER_BOTTOM = Border(
    bottom=Side(
        border_style=BORDER_THICK, color=Color('FF000000')
    )
)

FILL_GRAY = PatternFill(patternType='solid', fgColor=Color('D1D1D1'))
FILL_LIGHT_GRAY = PatternFill(patternType='solid', fgColor=Color('E1E1E1'))
FILL_LIGHTER_GRAY = PatternFill(patternType='solid', fgColor=Color('F1F1F1'))

FONT_RED = Font(color=RED)
FONT_GREEN = Font(color=GREEN)
FONT_HYPERLINK = Font(underline='single', color='0563C1')

HeaderLine = namedtuple('HeaderLine', ['egy_acronym', 'egy_title', 'code_header', 'title_header', 'credits_header',
                                       'block_header', 'mandatory_header'])
OfficialTextLine = namedtuple('OfficialTextLine', ['text'])
LearningUnitYearLine = namedtuple('LearningUnitYearLine', ['luy_acronym', 'luy_title'])
PrerequisiteItemLine = namedtuple(
    'PrerequisiteItemLine',
    ['text', 'operator', 'luy_acronym', 'luy_title', 'credits', 'block', 'mandatory']
)
PrerequisiteOfItemLine = namedtuple(
    'PrerequisitedItemLine',
    ['text', 'luy_acronym', 'luy_title', 'credits', 'block', 'mandatory']

)
HeaderLinePrerequisiteOf = namedtuple('HeaderLinePrerequisiteOf', ['egy_acronym', 'egy_title', 'title_header',
                                                                   'credits_header', 'block_header',
                                                                   'mandatory_header'])


class EducationGroupYearLearningUnitsPrerequisitesToExcel:
    def __init__(self, year: int, code: str):
        self.tree = ProgramTreeRepository.get(ProgramTreeIdentity(code, year))

    def _to_workbook(self):
        return generate_prerequisites_workbook(self.tree)

    def to_excel(self):
        return {
            'workbook': save_virtual_workbook(self._to_workbook()),
            'acronym': self.tree.root_node.title
        }


def generate_prerequisites_workbook(tree: 'ProgramTree') -> Workbook:
    worksheet_title = _("prerequisites-%(year)s-%(acronym)s") % {"year": tree.root_node.year,
                                                                 "acronym": tree.root_node.code}
    worksheet_title = clean_worksheet_title(worksheet_title)
    workbook = Workbook(encoding='utf-8')

    excel_lines = _build_excel_lines(tree)

    return _get_workbook(tree, excel_lines, workbook, worksheet_title, 7)


def _build_excel_lines(tree: 'ProgramTree') -> List:
    content = _first_line_content(
        HeaderLine(
            egy_acronym=tree.root_node.title,
            egy_title=tree.root_node.group_title_en if translation.get_language() == LANGUAGE_CODE_EN else
            tree.root_node.group_title_fr,
            code_header=_('Code'),
            title_header=_('Title'),
            credits_header=_('Cred. rel./abs.'),
            block_header=_('Block'),
            mandatory_header=_('Mandatory')
        )
    )

    for node in tree.get_nodes_that_have_prerequisites():
        content.append(
            LearningUnitYearLine(luy_acronym=node.code, luy_title=complete_title(node))
        )

        for group_number, group in enumerate(node.prerequisite.prerequisite_item_groups, start=1):
            for position, prerequisite_item in enumerate(group.prerequisite_items, start=1):
                prerequisite_item_links = tree.get_links_using_node(
                    tree.get_node_by_code_and_year(code=prerequisite_item.code, year=prerequisite_item.year)
                )
                prerequisite_line = _prerequisite_item_line(tree,
                                                            prerequisite_item, prerequisite_item_links,
                                                            node.prerequisite, group_number, position,
                                                            len(group.prerequisite_items))
                content.append(prerequisite_line)

    return content


def _first_line_content(header_line):
    content = list()
    content.append(
        header_line
    )
    content.append(
        OfficialTextLine(text=_("Official"))
    )
    return content


def _get_operator(prerequisite: Prerequisite, group_number: int, position: int):
    if group_number == 1 and position == 1:
        return None
    elif position == 1:
        return _(prerequisite.main_operator)
    return _(prerequisite.secondary_operator())


def _get_item_code(prerequisite_item: PrerequisiteItem, position: int, group_len: int):
    acronym_format = "{acronym}"

    if position == 1 and group_len > 1:
        acronym_format = "({acronym}"
    elif position == group_len and group_len > 1:
        acronym_format = "{acronym})"
    return acronym_format.format(acronym=prerequisite_item.code)


def _get_fill_to_apply(excel_lines: list):
    style_to_apply_dict = defaultdict(list)
    last_luy_line_index = None
    for index, row in enumerate(excel_lines, 1):
        if isinstance(row, HeaderLine):
            style_to_apply_dict[FILL_NO_GRAY].append("A{index}".format(index=index))
            style_to_apply_dict[FILL_NO_GRAY].append("B{index}".format(index=index))
            style_to_apply_dict[FILL_NO_GRAY].append("C{index}".format(index=index))
            style_to_apply_dict[FILL_NO_GRAY].append("D{index}".format(index=index))
            style_to_apply_dict[FILL_NO_GRAY].append("E{index}".format(index=index))
            style_to_apply_dict[FILL_NO_GRAY].append("F{index}".format(index=index))
            style_to_apply_dict[FILL_NO_GRAY].append("G{index}".format(index=index))

        elif isinstance(row, LearningUnitYearLine):
            style_to_apply_dict[FILL_GRAY].append("A{index}".format(index=index))
            style_to_apply_dict[FILL_LIGHT_GRAY].append("B{index}".format(index=index))
            last_luy_line_index = index

        elif isinstance(row, PrerequisiteItemLine):
            if (last_luy_line_index - index) % 2 == 1:
                style_to_apply_dict[FILL_LIGHTER_GRAY].append("C{index}".format(index=index))
                style_to_apply_dict[FILL_LIGHTER_GRAY].append("D{index}".format(index=index))
                style_to_apply_dict[FILL_LIGHTER_GRAY].append("E{index}".format(index=index))
                style_to_apply_dict[FILL_LIGHTER_GRAY].append("F{index}".format(index=index))
                style_to_apply_dict[FILL_LIGHTER_GRAY].append("G{index}".format(index=index))
        elif isinstance(row, PrerequisiteOfItemLine):
            if (last_luy_line_index - index) % 2 == 1:
                style_to_apply_dict[FILL_LIGHTER_GRAY].append("C{index}".format(index=index))
                style_to_apply_dict[FILL_LIGHTER_GRAY].append("D{index}".format(index=index))
                style_to_apply_dict[FILL_LIGHTER_GRAY].append("E{index}".format(index=index))
                style_to_apply_dict[FILL_LIGHTER_GRAY].append("F{index}".format(index=index))
        if isinstance(row, HeaderLinePrerequisiteOf):
            style_to_apply_dict[FILL_NO_GRAY].append("A{index}".format(index=index))
            style_to_apply_dict[FILL_NO_GRAY].append("B{index}".format(index=index))
            style_to_apply_dict[FILL_NO_GRAY].append("C{index}".format(index=index))
            style_to_apply_dict[FILL_NO_GRAY].append("D{index}".format(index=index))
            style_to_apply_dict[FILL_NO_GRAY].append("E{index}".format(index=index))
            style_to_apply_dict[FILL_NO_GRAY].append("F{index}".format(index=index))
    return style_to_apply_dict


def _get_font_to_apply(excel_lines: list):
    font_to_apply_dict = defaultdict(list)
    for index, row in enumerate(excel_lines, 1):
        if isinstance(row, PrerequisiteItemLine):
            if row.operator == _(OR):
                font_to_apply_dict[FONT_RED].append("B{index}".format(index=index))
            elif row.operator == _(AND):
                font_to_apply_dict[FONT_GREEN].append("B{index}".format(index=index))
    return font_to_apply_dict


def _get_border_to_apply(excel_lines: list):
    border_to_apply_dict = defaultdict(list)
    for index, row in enumerate(excel_lines, 1):
        if isinstance(row, OfficialTextLine):
            border_to_apply_dict[BORDER_BOTTOM].append("A{index}".format(index=index))
    return border_to_apply_dict


def _merge_cells(excel_lines, workbook: Workbook, end_column):
    worksheet = workbook.worksheets[0]
    for index, row in enumerate(excel_lines, 1):
        if isinstance(row, LearningUnitYearLine):
            worksheet.merge_cells(start_row=index, end_row=index, start_column=2, end_column=end_column)


def _add_hyperlink(excel_lines, workbook: Workbook, year):
    worksheet = workbook.worksheets[0]
    for index, row in enumerate(excel_lines, 1):
        if isinstance(row, LearningUnitYearLine):
            cell = worksheet.cell(row=index, column=1)
            cell.hyperlink = LEARNING_UNIT_PORTAL_URL.format(year=year, acronym=row.luy_acronym)
            cell.font = FONT_HYPERLINK

        if isinstance(row, PrerequisiteItemLine) or isinstance(row, PrerequisiteOfItemLine):
            column_nb = 3 if isinstance(row, PrerequisiteItemLine) else 2
            cell = worksheet.cell(row=index, column=column_nb)
            cell.hyperlink = LEARNING_UNIT_PORTAL_URL.format(year=year, acronym=row.luy_acronym.strip("()"))
            cell.font = FONT_HYPERLINK


class EducationGroupYearLearningUnitsIsPrerequisiteOfToExcel:

    def __init__(self, year: int, code: str):
        self.tree = ProgramTreeRepository.get(ProgramTreeIdentity(code, year))
        self.acronym = "{}-{}"

    def _to_workbook(self):
        return generate_ue_is_prerequisite_for_workbook(self.tree)

    def to_excel(self):
        return {
            'workbook': save_virtual_workbook(self._to_workbook()),
            'acronym': self.tree.root_node.title
        }


def generate_ue_is_prerequisite_for_workbook(tree: 'ProgramTree') -> Workbook:
    worksheet_title = _("is_prerequisite_of-%(year)s-%(acronym)s") % {"year": tree.root_node.year,
                                                                      "acronym": tree.root_node.code}
    worksheet_title = clean_worksheet_title(worksheet_title)
    workbook = Workbook()

    excel_lines = _build_excel_lines_prerequisited(tree)
    return _get_workbook(tree, excel_lines, workbook, worksheet_title, 6)


def _get_workbook(tree: 'ProgramTree',
                  excel_lines: List,
                  workbook: Workbook,
                  worksheet_title: str,
                  end_column: int) -> Workbook:
    header, *content = [tuple(line) for line in excel_lines]
    fill = _get_fill_to_apply(excel_lines)
    fonts = _get_font_to_apply(excel_lines)
    worksheet_data = {
        WORKSHEET_TITLE_KEY: worksheet_title,
        HEADER_TITLES_KEY: header,
        CONTENT_KEY: content,
        FILL_CELLS: fill,
        FONT_CELLS: fonts
    }
    _build_worksheet(worksheet_data, workbook, 0)
    _merge_cells(excel_lines, workbook, end_column)
    _add_hyperlink(excel_lines, workbook, str(tree.root_node.year))
    return workbook


def _build_excel_lines_prerequisited(tree: 'ProgramTree') -> List:
    content = _first_line_content(HeaderLinePrerequisiteOf(egy_acronym=tree.root_node.code,
                                                           egy_title=tree.root_node.title,
                                                           title_header=_('Title'),
                                                           credits_header=_('Cred. rel./abs.'),
                                                           block_header=_('Block'),
                                                           mandatory_header=_('Mandatory'))
                                  )
    for child_node in tree.get_nodes_that_are_prerequisites():
        if child_node.is_prerequisite:
            content.append(
                LearningUnitYearLine(luy_acronym=child_node.code, luy_title=child_node.title)
            )
            first = True
            for prerequisite_node in child_node.get_is_prerequisite_of():
                if child_node.year == prerequisite_node.year:
                    prerequisite_line = _build_is_prerequisite_for_line(
                        prerequisite_node,
                        first,
                        tree
                    )
                    first = False
                    content.append(prerequisite_line)
    return content


def _build_is_prerequisite_for_line(prerequisite_node: 'NodeLearningUnitYear', first, tree: 'ProgramTree') \
        -> PrerequisiteOfItemLine:
    text = (_("is a prerequisite of") + " :") if first else None
    first_link = tree.get_first_link_occurence_using_node(prerequisite_node)
    return PrerequisiteOfItemLine(
        text=text,
        luy_acronym=prerequisite_node.code,
        luy_title=prerequisite_node.title,
        credits=first_link.relative_credits_repr,
        block=first_link.block_repr,
        mandatory=_("Yes") if first_link.is_mandatory else _("No")
    )


def clean_worksheet_title(title: str) -> str:
    # Worksheet title is max 25 chars (31 chars with sheet number) + does not accept slash present in acronyms
    return title[:25].replace("/", "_")


def _prerequisite_item_line(tree, prerequisite_item: PrerequisiteItem, links: List[link.Link],
                            prerequisite: Prerequisite, group_number: int, position: int,
                            number_of_prerequisite_item: int):
    item_link = links[0]

    text = (_("has as prerequisite") + " :") \
        if group_number == 1 and position == 1 else None

    return PrerequisiteItemLine(
        text=text,
        operator=_get_operator(prerequisite, group_number, position),
        luy_acronym=_get_item_code(prerequisite_item, position, number_of_prerequisite_item),
        luy_title=item_link.child.title,
        credits=tree.get_relative_credits_values(NodeIdentity(prerequisite_item.code, prerequisite_item.year)),
        block=tree.get_blocks_values(NodeIdentity(prerequisite_item.code, prerequisite_item.year)),
        mandatory=_("Yes") if item_link and item_link.is_mandatory else _("No")
    )


def complete_title(luy: 'NodeLearningUnitYear'):
    if translation.get_language() == LANGUAGE_CODE_EN:
        return luy.full_title_en
    return luy.full_title_fr
